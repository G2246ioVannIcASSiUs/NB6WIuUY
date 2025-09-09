# 代码生成时间: 2025-09-09 19:26:03
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import View
from django.urls import path
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime

def generate_excel_response(data, title):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Generated Excel'
    ws.append(['Title 1', 'Title 2', 'Title 3'])  # Example headers
    for row in data:
        ws.append(row)
    
    # Set the title of the Excel file
    wb.save(title + '.xlsx')
    wb = openpyxl.load_workbook(title + '.xlsx')
    wb.remove(wb['Sheet'])
    response = HttpResponse(wb, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment;filename="{title}.xlsx"'
    return response

def generate_excel_file(data, title):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Generated Excel'
    ws.append(['Title 1', 'Title 2', 'Title 3'])  # Example headers
    for row in data:
        ws.append(row)
    
    # Save the Excel file
    wb.save(title + '.xlsx')

def generate_excel_view(request):
    """
    View to generate an Excel file based on provided data.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: A response containing the generated Excel file.
    """
    if request.method == 'GET':
        # Example data for generating the Excel file
        data = [
            ['Row 1, Column 1', 'Row 1, Column 2', 'Row 1, Column 3'],
            ['Row 2, Column 1', 'Row 2, Column 2', 'Row 2, Column 3'],
            # More rows can be added here
        ]
        title = 'Generated_Excel_File'
        return generate_excel_response(data, title)
    else:
        return HttpResponse('Invalid request', status=400)

def generate_excel_file_view(request):
    """
    View to generate and save an Excel file based on provided data.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: A response indicating the file has been saved.
    """
    if request.method == 'GET':
        # Example data for generating the Excel file
        data = [
            ['Row 1, Column 1', 'Row 1, Column 2', 'Row 1, Column 3'],
            ['Row 2, Column 1', 'Row 2, Column 2', 'Row 2, Column 3'],
            # More rows can be added here
        ]
        title = 'Generated_Excel_File'
        generate_excel_file(data, title)
        return HttpResponse(f'Excel file {title}.xlsx has been saved.', status=200)
    else:
        return HttpResponse('Invalid request', status=400)
